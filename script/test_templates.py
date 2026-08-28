"""Проверка круга: правка шаблона → скачивание → заполнение → повторная загрузка."""
import json, io, sys, urllib.request, urllib.parse
import openpyxl

BASE = "http://localhost:5000"
TOKEN = open("/tmp/tok").read().strip()


def req(method, path, data=None, raw=False, files=None):
    url = BASE + path
    headers = {"x-auth-token": TOKEN}
    body = None
    if files:
        boundary = "----grrtest"
        name, fname, content = files
        pre = (f"--{boundary}\r\nContent-Disposition: form-data; name=\"{name}\"; filename=\"{fname}\"\r\n"
               "Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n").encode()
        body = pre + content + f"\r\n--{boundary}--\r\n".encode()
        headers["Content-Type"] = f"multipart/form-data; boundary={boundary}"
    elif data is not None:
        body = json.dumps(data).encode()
        headers["Content-Type"] = "application/json"
    r = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(r) as resp:
            b = resp.read()
            return resp.status, (b if raw else json.loads(b))
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode()


steps = []

# 0. заводской вид, чтобы тест был повторяемым
req("POST", "/api/templates/fuel/reset")

# 1. читаем шаблон ГСМ
st, d = req("GET", "/api/templates/fuel")
assert st == 200, d
cols = d["def"]["columns"]
steps.append(("шаблон fuel прочитан", [c["label"] for c in cols]))

# 2. переименовываем колонки, меняем порядок, добавляем свою, задаём лист и подсказки
labels = {"date": "Дата заправки", "object": "Участок работ", "unit": "Единица техники (борт. №)",
          "norm": "Норма по паспорту, л", "fact": "Выдано фактически, л",
          "liters": "Литры выдано", "hours": "Моточасы", "meters": "Проходка, м"}
new_cols = []
for c in cols:
    c = dict(c)
    if c["key"] in labels:
        c["label"] = labels[c["key"]]
    c["hint"] = "дд.мм.гггг" if c["key"] == "date" else c.get("hint", "")
    new_cols.append(c)
new_cols = [new_cols[1], new_cols[0]] + new_cols[2:]           # порядок: станок, дата, ...
new_cols.append({"key": "own_note", "label": "Мой комментарий", "hint": "любой текст",
                 "required": False, "custom": True})
payload = {"title": "Журнал ГСМ (моя форма)", "sheetName": "ГСМ участок",
           "columns": new_cols, "notes": ["Заполняется механиком участка."],
           "baseType": "fuel"}
st, d = req("PUT", "/api/templates/fuel", payload)
assert st == 200, d
steps.append(("шаблон сохранён", [c["label"] for c in d["def"]["columns"]]))

# 3. скачиваем изменённый шаблон
st, xlsx = req("GET", "/api/templates/fuel/xlsx", raw=True)
assert st == 200 and xlsx[:2] == b"PK", st
open("/tmp/tpl_fuel.xlsx", "wb").write(xlsx)
wb = openpyxl.load_workbook(io.BytesIO(xlsx))
ws = wb[wb.sheetnames[0]]
rows = [[c.value for c in r] for r in ws.iter_rows(max_row=8)]
steps.append(("скачан файл", {"лист": ws.title, "строки": rows[:5]}))

# ищем строку шапки
hdr_i = next(i for i, r in enumerate(rows)
             if r and any(str(v or "").startswith("Участок работ") for v in r))
headers = [str(v or "") for v in rows[hdr_i]]

# 4. чистим примеры и вписываем свои данные под новыми названиями
ws.delete_rows(hdr_i + 2, ws.max_row - hdr_i)   # всё после шапки
# берём реальный объект из справочника и строим строки в порядке изменённых колонок
st, sett = req("GET", "/api/settings")
objs = sett.get("objects", []) if isinstance(sett, dict) else []
obj = objs[0]["name"] if objs else "Участок «Северный»"
data = [
    {"Участок работ *": obj, "Дата заправки *": "05.08.2026", "Единица техники *": "УКБ-1",
     "Норма, л *": 300, "Факт, л *": 320, "Мой комментарий": "ночная смена"},
    {"Участок работ *": obj, "Дата заправки *": "06.08.2026", "Единица техники *": "УКБ-1",
     "Норма, л *": 300, "Факт, л *": 280, "Мой комментарий": ""},
]
data = [[row.get(h, "") for h in headers] for row in data]
for r in data:
    ws.append(r)
buf = io.BytesIO()
wb.save(buf)
filled = buf.getvalue()
open("/tmp/tpl_fuel_filled.xlsx", "wb").write(filled)

# 5. загружаем обратно
st, up = req("POST", "/api/import/upload", files=("file", "ГСМ участок.xlsx", filled))
assert st == 200, up
print("ЗАГРУЗКА:", json.dumps(up, ensure_ascii=False)[:1500])
steps.append(("распознано при загрузке", up.get("suggestedType") or up))

# 6. предпросмотр строк с изменёнными названиями колонок
if "uploadId" in up:
    t = up.get("suggestedType") or "fuel"
    st, mp = req("POST", "/api/import/mapping", {"headers": up.get("headers", []), "type": t})
    steps.append(("сопоставление колонок", mp))
    st, pv = req("POST", "/api/import/preview",
                 {"uploadId": up["uploadId"], "type": t, "mapping": mp["mapping"]})
    steps.append(("предпросмотр", {k: v for k, v in pv.items() if k != "items"}))
    print("ПРЕДПРОСМОТР:", json.dumps(pv, ensure_ascii=False)[:1200])

for s in steps:
    print("\n---", s[0])
    print(json.dumps(s[1], ensure_ascii=False)[:800])
