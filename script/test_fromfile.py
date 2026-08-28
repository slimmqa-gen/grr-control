"""Проверка: шаблон по своему файлу + брендирование в выгрузке Excel."""
import json, io, urllib.request, urllib.error, os
import openpyxl

BASE = "http://localhost:5000"
TOKEN = open("/tmp/tok").read().strip()


def req(method, path, data=None, raw=False, files=None):
    headers = {"x-auth-token": TOKEN}
    body = None
    if files:
        b = "----grrtest"
        name, fname, content, ctype = files
        body = (f"--{b}\r\nContent-Disposition: form-data; name=\"{name}\"; filename=\"{fname}\"\r\n"
                f"Content-Type: {ctype}\r\n\r\n").encode() + content + f"\r\n--{b}--\r\n".encode()
        headers["Content-Type"] = f"multipart/form-data; boundary={b}"
    elif data is not None:
        body = json.dumps(data).encode()
        headers["Content-Type"] = "application/json"
    r = urllib.request.Request(BASE + path, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(r) as resp:
            raw_b = resp.read()
            return resp.status, (raw_b if raw else json.loads(raw_b))
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode()


# 1. шаблон на основе рабочего файла заказчика
src = "pbk_files/Svodka-uch-k-Lineinyi.xlsx"
content = open(src, "rb").read()
st, prop = req("POST", "/api/templates/from-file",
               files=("file", os.path.basename(src), content,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))
print("from-file:", st)
print(json.dumps({k: v for k, v in prop.items() if k != "sampleRows"}, ensure_ascii=False)[:900])

# 2. сохраняем как свой шаблон и профиль
st, created = req("POST", "/api/templates", {
    "title": prop["title"], "sheetName": prop["sheetName"],
    "columns": prop["columns"], "notes": prop.get("notes", []),
    "baseType": prop.get("baseType", ""),
})
print("создан:", st, created["def"]["code"] if st == 200 else created)
code = created["def"]["code"] if st == 200 else None

st, tl = req("GET", "/api/templates")
print("всего шаблонов:", len(tl["templates"]))
st, profs = req("GET", "/api/profiles")
names = [p.get("name") for p in (profs.get("profiles") if isinstance(profs, dict) else profs or [])]
print("профили:", names)

# 3. скачиваем созданный шаблон
if code:
    st, x = req("GET", f"/api/templates/{code}/xlsx", raw=True)
    wb = openpyxl.load_workbook(io.BytesIO(x))
    ws = wb[wb.sheetnames[0]]
    print("лист:", ws.title)
    for i, r in enumerate(ws.iter_rows(max_row=6, values_only=True), 1):
        print(" ", i, [str(v)[:22] if v is not None else "" for v in r][:8])

# 4. брендирование
st, b = req("PUT", "/api/branding", {
    "orgName": "ООО «Производственно-Буровая Компания»",
    "orgShort": "ПБК",
    "orgInn": "2460000000 / 246001001",
    "orgDetails": "660000, г. Красноярск, ул. Ленина, 1 · тел. +7 391 000-00-00",
    "signerName": "Петрова А. С.",
    "signerPosition": "аналитик",
    "logo": "",
})
print("брендирование:", st, json.dumps(b, ensure_ascii=False)[:200] if st == 200 else b)

# 5. брендированная выгрузка
st, x = req("GET", "/api/export/fuel", raw=True)
print("выгрузка ГСМ:", st, len(x))
wb = openpyxl.load_workbook(io.BytesIO(x))
ws = wb[wb.sheetnames[0]]
for i, r in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
    print(" ", i, [str(v)[:40] if v is not None else "" for v in r][:5])
last = [c.value for c in list(ws.iter_rows(min_row=ws.max_row))[0]]
print("подпись:", str(last[0])[:80])

# 6. шаблон с брендом в шапке
st, x = req("GET", "/api/templates/reports/xlsx", raw=True)
ws = openpyxl.load_workbook(io.BytesIO(x))[openpyxl.load_workbook(io.BytesIO(x)).sheetnames[0]]
print("шапка шаблона:", str(ws.cell(1, 1).value)[:110])
